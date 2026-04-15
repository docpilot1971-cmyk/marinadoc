"""Integration test: parse contract.docx and generate KS-2/KS-3, compare with reference files."""
from __future__ import annotations

import shutil
from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import load_workbook

from app.core.config import AppConfig, PathsConfig, TemplatesConfig
from app.core.logging_setup import setup_logging
from app.models import PartyType, RowGroupingMode
from app.services.excel_template_processor import ExcelTemplateProcessor
from app.services.stubs.classifier_stub import ContractTypeClassifierStub
from app.services.stubs.generators_stub import KS2ExcelGeneratorStub, KS3ExcelGeneratorStub
from app.services.stubs.parsers_stub import (
    HeaderParserStub,
    ObjectParserStub,
    PartiesParserStub,
    PeriodParserStub,
    TableParserStub,
    TotalsParserStub,
)
from app.services.stubs.reader_stub import ContractReaderStub
from app.services.stubs.validator_stub import ExtractionValidatorStub
from app.services.template_loader import TemplateLoader
from app.services.word_template_processor import WordTemplateProcessor

PROJECT_ROOT = Path(__file__).resolve().parent.parent
INCOMING = PROJECT_ROOT / "templates" / "incoming"
CONTRACT_PATH = INCOMING / "contract.docx"
KS2_REFERENCE = INCOMING / "ks2_verno.xlsx"
KS3_REFERENCE = INCOMING / "ks3_verno.xlsx"


@pytest.fixture(scope="module", autouse=True)
def _init_logging():
    setup_logging()


def _prepare_output_dir() -> Path:
    out = PROJECT_ROOT / "output" / "test_comparison"
    if out.exists():
        shutil.rmtree(out)
    out.mkdir(parents=True, exist_ok=True)
    return out


def _build_controller():
    """Build minimal controller with real parser stubs."""
    config = AppConfig(
        paths=PathsConfig(
            templates_dir=INCOMING,
            output_dir=PROJECT_ROOT / "output",
            preview_dir=PROJECT_ROOT / "output" / "preview",
        ),
        templates=TemplatesConfig(
            act_word_template="act_org_org_filled_test.docx",
            act_word_template_ip="act_org_ip_filled_test.docx",
            ks2_template="ks2_test.xlsx",
            ks3_template="ks3_test.xlsx",
        ),
    )
    template_loader = TemplateLoader(config)
    word_processor = WordTemplateProcessor()
    excel_processor = ExcelTemplateProcessor()

    reader = ContractReaderStub()
    classifier = ContractTypeClassifierStub()
    header_parser = HeaderParserStub()
    parties_parser = PartiesParserStub()
    object_parser = ObjectParserStub()
    period_parser = PeriodParserStub()
    table_parser = TableParserStub()
    totals_parser = TotalsParserStub()
    validator = ExtractionValidatorStub()
    ks2_gen = KS2ExcelGeneratorStub(template_loader, excel_processor)
    ks3_gen = KS3ExcelGeneratorStub(template_loader, excel_processor)

    return {
        "config": config,
        "reader": reader,
        "classifier": classifier,
        "header_parser": header_parser,
        "parties_parser": parties_parser,
        "object_parser": object_parser,
        "period_parser": period_parser,
        "table_parser": table_parser,
        "totals_parser": totals_parser,
        "validator": validator,
        "ks2_gen": ks2_gen,
        "ks3_gen": ks3_gen,
    }


def _full_extraction(contract_path: Path, ctrl: dict):
    """Run the full parsing pipeline on a contract file."""
    contract = ctrl["reader"].read(contract_path)
    cls = ctrl["classifier"].classify(contract)
    doc_data = ctrl["header_parser"].parse(contract)
    customer, executor = ctrl["parties_parser"].parse(contract)
    obj_data = ctrl["object_parser"].parse(contract)
    period_data = ctrl["period_parser"].parse(contract)
    rows = ctrl["table_parser"].parse(contract, cls.table_grouping_mode)
    totals = ctrl["totals_parser"].parse(contract, rows)

    from app.models import ExtractionResult

    result = ExtractionResult(
        document=doc_data,
        customer=customer,
        executor=executor,
        object_data=obj_data,
        period=period_data,
        rows=rows,
        totals=totals,
    )
    # Run validator
    ctrl["validator"].validate(result)
    return contract, result


class TestContractParsingAndGeneration:
    """Parse contract.docx and generate KS-2/KS-3, then compare with reference files."""

    @pytest.fixture(autouse=True)
    def setup(self):
        self.output_dir = _prepare_output_dir()
        self.ctrl = _build_controller()
        assert CONTRACT_PATH.exists(), f"Contract file not found: {CONTRACT_PATH}"
        assert KS2_REFERENCE.exists(), f"KS-2 reference not found: {KS2_REFERENCE}"
        assert KS3_REFERENCE.exists(), f"KS-3 reference not found: {KS3_REFERENCE}"

        self.contract, self.result = _full_extraction(CONTRACT_PATH, self.ctrl)

    def test_contract_parses_successfully(self):
        assert self.contract is not None
        assert len(self.contract.paragraphs) > 0
        assert len(self.contract.tables) > 0

    def test_header_extraction(self):
        doc = self.result.document
        assert doc.contract_number is not None, "Contract number not extracted"
        assert doc.contract_date is not None, "Contract date not extracted"

    def test_parties_extraction(self):
        assert self.result.customer.full_name is not None, "Customer name not extracted"
        assert self.result.executor.full_name is not None, "Executor name not extracted"
        assert self.result.customer.inn is not None, "Customer INN not extracted"
        assert self.result.executor.inn is not None, "Executor INN not extracted"

    def test_rows_extracted(self):
        assert len(self.result.rows) > 0, "No estimate rows extracted from contract"

    def test_totals_extracted(self):
        t = self.result.totals
        assert t.total_without_vat > 0, "Total without VAT not extracted"
        assert t.total_with_vat > 0, "Total with VAT not extracted"

    def test_ks2_generation(self):
        output_path = self.output_dir / "ks2_generated.xlsx"
        generated = self.ctrl["ks2_gen"].generate(self.result, output_path)
        assert generated.exists(), "KS-2 file was not generated"
        self._compare_excel(generated, KS2_REFERENCE, "KS-2")

    def test_ks3_generation(self):
        output_path = self.output_dir / "ks3_generated.xlsx"
        generated = self.ctrl["ks3_gen"].generate(self.result, output_path)
        assert generated.exists(), "KS-3 file was not generated"
        self._compare_excel(generated, KS3_REFERENCE, "KS-3")

    def _compare_excel(self, generated_path: Path, reference_path: Path, label: str):
        """Compare two Excel files cell-by-cell, reporting differences."""
        wb_gen = load_workbook(generated_path, data_only=False)
        wb_ref = load_workbook(reference_path, data_only=False)

        ws_gen = wb_gen.active
        ws_ref = wb_ref.active

        diffs = []
        max_row = max(ws_gen.max_row, ws_ref.max_row)
        max_col = max(ws_gen.max_column, ws_ref.max_column)

        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                cell_gen = ws_gen.cell(row, col).value
                cell_ref = ws_ref.cell(row, col).value

                # Normalize for comparison
                gen_str = self._normalize(cell_gen)
                ref_str = self._normalize(cell_ref)

                if gen_str != ref_str and (cell_ref is not None or cell_gen is not None):
                    if cell_ref is not None:  # Only report if reference has content
                        coord = f"{chr(64 + col)}{row}" if col <= 26 else f"{'A' * (col // 26)}{chr(64 + col % 26)}{row}"
                        diffs.append(
                            f"  {coord}: ref={cell_ref!r}, gen={cell_gen!r}"
                        )

        if diffs:
            msg = f"\n{label} differences ({len(diffs)}):\n" + "\n".join(diffs[:50])
            if len(diffs) > 50:
                msg += f"\n... and {len(diffs) - 50} more"
            pytest.fail(msg)

    @staticmethod
    def _normalize(value):
        if value is None:
            return None
        if isinstance(value, str):
            return value.strip().replace("\xa0", " ").replace("\r\n", "\n")
        if isinstance(value, Decimal):
            return float(value)
        return value
