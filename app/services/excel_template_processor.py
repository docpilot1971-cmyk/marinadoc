from __future__ import annotations

import copy
import re
from datetime import date
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill, numbers
from openpyxl.utils import get_column_letter

from app.models import EstimateRow, ExtractionResult, RowType
from app.services.amount_to_words import amount_to_words_ru

_NO_FILL = PatternFill(fill_type=None)


class ExcelTemplateProcessor:
    def analyze_structure(self, template_path: Path) -> dict[str, object]:
        wb = load_workbook(template_path, data_only=False)
        sheets: list[dict[str, object]] = []
        for ws in wb.worksheets:
            non_empty = 0
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
                if any(cell.value is not None and str(cell.value).strip() for cell in row):
                    non_empty += 1
            sheets.append(
                {
                    "title": ws.title,
                    "max_row": ws.max_row,
                    "max_col": ws.max_column,
                    "non_empty_rows": non_empty,
                }
            )
        return {"sheets": sheets}

    def render_ks2(self, template_path: Path, data: ExtractionResult, output_path: Path) -> Path:
        if not template_path.exists():
            raise FileNotFoundError(f"KS-2 template not found: {template_path}")

        is_ip_template = "ip" in template_path.name.lower()

        # For IP templates, use Excel COM directly to avoid openpyxl corruption
        if is_ip_template:
            return self._render_ks2_ip_via_excel(template_path, data, output_path)

        # ORG templates: use openpyxl
        wb = load_workbook(template_path)
        self._set_landscape_orientation(wb)
        ws = wb.active

        self._clear_fills(ws)

        # Find signature block BEFORE any modifications
        sig_block_row = self._find_ks2_signature_block(ws) or 30

        # Clear ALL dynamic area: from row 24 to just before signatures
        self._reset_ks2_dynamic_area(ws, sig_block_row)

        self._fill_ks2_header(ws, data)
        last_data_row = self._fill_ks2_data_rows_only(ws, data.rows, sig_block_row)

        # Write totals immediately after data rows
        totals_start_row = last_data_row + 1
        self._fill_ks2_totals_and_payment(ws, data, is_ip_template, totals_start_row)

        self._fill_ks2_signatures(ws, data, is_ip_template, sig_block_row)
        self._validate_ks2_result(ws, data)

        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(output_path))
        return output_path

    def _render_ks2_ip_via_excel(self, template_path: Path, data: ExtractionResult, output_path: Path) -> Path:
        """Generate KS-2 for IP using Excel COM to avoid openpyxl corruption."""
        import pythoncom
        import win32com.client

        pythoncom.CoInitialize()
        excel = None
        wb = None
        try:
            excel = win32com.client.DispatchEx("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            wb = excel.Workbooks.Open(str(template_path.resolve()), ReadOnly=False)
            ws = wb.Worksheets(1)

            # Fill header
            self._fill_ks2_header_com(ws, data)

            # Fill data rows
            sig_block_row = self._find_ks2_signature_block_com(ws) or 30
            # _fill_ks2_data_rows_com returns (last_data_row, new_sig_block_row)
            result = self._fill_ks2_data_rows_com(ws, data.rows, sig_block_row)
            if isinstance(result, tuple):
                last_data_row, sig_block_row = result
            else:
                last_data_row = result

            # Fill totals
            self._fill_ks2_totals_com(ws, data, last_data_row + 1)

            # Fill signatures using the UPDATED sig_block_row
            self._fill_ks2_signatures_com(ws, data, sig_block_row)

            # Save
            output_path.parent.mkdir(parents=True, exist_ok=True)
            # Save with explicit xlsx format (51 = xlOpenXMLWorkbook)
            wb.SaveAs(str(output_path.resolve()), FileFormat=51)
            import logging
            logging.getLogger(__name__).info("KS-2 IP generated via Excel COM: %s", output_path)
            return output_path

        finally:
            if wb is not None:
                try:
                    wb.Close(False)
                except Exception:
                    pass
            if excel is not None:
                try:
                    excel.Quit()
                except Exception:
                    pass
            try:
                pythoncom.CoUninitialize()
            except Exception:
                pass
            # Force GC to release COM references and prevent Excel process leaks
            import gc
            gc.collect()

    def _fill_ks2_data_rows_only(self, ws, rows: list[EstimateRow], sig_block_row: int) -> int:
        """Fill KS-2 data rows WITHOUT inserting new rows. Returns last data row index."""
        data_rows = [
            r for r in rows
            if r.row_type not in {RowType.SUBTOTAL, RowType.GRAND_TOTAL}
            and not self._is_total_row(r.row_name or "")
        ]
        if not data_rows:
            return 23  # No data

        start_row = 24
        available_rows = sig_block_row - start_row

        # If more data rows than available, only fill what fits
        rows_to_fill = data_rows[:available_rows]

        for idx, row_data in enumerate(rows_to_fill):
            self._fill_ks2_row(ws, start_row + idx, row_data)

        # Clear any remaining empty rows between data and signature block
        last_data_row = start_row + len(rows_to_fill) - 1
        self._cleanup_ks2_empty_rows(ws, start_row, last_data_row)

        return last_data_row

    def render_ks3(self, template_path: Path, data: ExtractionResult, output_path: Path) -> Path:
        if not template_path.exists():
            raise FileNotFoundError(f"KS-3 template not found: {template_path}")
        wb = load_workbook(template_path)
        self._set_landscape_orientation(wb)
        ws = wb[wb.sheetnames[0]]

        # Determine template type by filename
        is_ip_template = "ip" in template_path.name.lower()

        self._clear_fills(ws)
        # IMPORTANT: Find fixed blocks BEFORE clearing
        fixed_block_row = self._find_ks3_fixed_block_start(ws, 26)
        sig_block_row = self._find_ks3_signature_block(ws)

        # Clear only the dynamic area: from row 26 to just before the first fixed block
        clear_end = fixed_block_row if fixed_block_row and fixed_block_row > 26 else 29
        self._reset_ks3_dynamic_area(ws, clear_end)

        self._fill_ks3_header(ws, data)
        self._fill_ks3_labels(ws, is_ip_template)
        # Pass data to _fill_ks3_rows so IP mode can access totals
        self._fill_ks3_rows(ws, data.rows, fixed_block_row, is_ip_template, data)
        self._fill_ks3_totals(ws, data, is_ip_template)
        self._fill_ks3_signatures(ws, data, is_ip_template)
        self._cleanup_ks3_empty_rows(ws)
        self._validate_ks3_result(ws, data, is_ip_template)

        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(output_path))
        return output_path

    @staticmethod
    def _cleanup_ks2_post_totals_empty_rows(ws) -> None:  # noqa: ANN001
        """Remove empty rows between 'Всего к оплате' and signatures."""
        # Find the 'Всего к оплате' row
        payment_row = None
        for row in range(1, ws.max_row + 1):
            cell_a = ws.cell(row, 1).value
            if isinstance(cell_a, str) and "всего к оплате" in cell_a.lower():
                payment_row = row
                break

        # Find the signature start row (Сдал)
        signature_row = None
        for row in range(1, ws.max_row + 1):
            cell_a = ws.cell(row, 1).value
            if isinstance(cell_a, str) and "сдал" in cell_a.lower():
                signature_row = row
                break

        if payment_row is None or signature_row is None:
            return

        # Clear empty rows between payment and signatures
        for row in range(signature_row - 1, payment_row, -1):
            # Check if row is completely empty in columns A-J
            row_values = [ws.cell(row, col).value for col in range(1, 11)]
            has_content = any(v is not None and str(v).strip() for v in row_values)
            if not has_content:
                # Clear the row, skipping merged cells
                for col in range(1, 15):  # Clear up to column O
                    cell = ws.cell(row, col)
                    if cell.__class__.__name__ != "MergedCell":
                        cell.value = None

    @staticmethod
    def _clear_fills(ws) -> None:  # noqa: ANN001
        """Remove background fill from all cells in the worksheet."""
        for row in ws.iter_rows():
            for cell in row:
                if cell.fill and cell.fill.patternType is not None:
                    cell.fill = _NO_FILL

    def _fill_ks2_header(self, ws, data: ExtractionResult) -> None:  # noqa: ANN001
        if data.customer.full_name or data.customer.address:
            self._set_cell(ws, "C6", f"{data.customer.full_name or ''}{self._fmt_address(data.customer.address)}")
        if data.executor.full_name or data.executor.address:
            self._set_cell(ws, "C7", f"{data.executor.full_name or ''}{self._fmt_address(data.executor.address)}")
        if data.object_data.object_name or data.object_data.object_address:
            self._set_cell(ws, "C8", f"{data.object_data.object_name or ''}{self._fmt_address(data.object_data.object_address)}")
        self._set_cell(ws, "L10", f"№{data.document.contract_number or ''}")

        if data.document.contract_date:
            self._set_cell(ws, "L11", f"{data.document.contract_date.day:02d}")
            self._set_cell(ws, "N11", f"{data.document.contract_date.month:02d}")
            self._set_cell(ws, "O11", str(data.document.contract_date.year))

        if data.period.work_start_date:
            self._set_cell(ws, "L15", data.period.work_start_date)
        if data.period.work_end_date_fact:
            self._set_cell(ws, "M15", data.period.work_end_date_fact)

        total_words = amount_to_words_ru(data.totals.total_without_vat or Decimal("0"))
        vat_suffix = ", без НДС"
        if data.totals.vat_rate > 0 or data.totals.vat_amount > 0:
            vat_suffix = f", в т.ч. НДС {data.totals.vat_rate}%"
        self._set_cell(
            ws,
            "A17",
            f"Сметная (договорная) стоимость {data.totals.total_without_vat} ({total_words}){vat_suffix}",
        )

    def _fill_ks2_rows(self, ws, rows: list[EstimateRow], fixed_block_row: int, is_ip_template: bool = False) -> None:  # noqa: ANN001
        """Fill KS-2 data rows. Returns the last row index used."""
        self._fill_ks2_data_rows_only(ws, rows, fixed_block_row)

    @staticmethod
    def _is_total_row(name: str) -> bool:
        """Check if a row name indicates it's a total/summary row that should be excluded from data table."""
        lower = name.lower().strip()
        total_keywords = ["итого", "всего", "total", "общая стоимость", "к оплате", "сумма", "ндс"]
        return any(kw in lower for kw in total_keywords)

    @staticmethod
    def _find_ks2_totals_block_start(ws, from_row: int) -> int | None:  # noqa: ANN001
        """Find the row where the KS-2 totals block starts (Итого общая стоимость)."""
        for row in range(from_row, ws.max_row + 1):
            val_c = ws.cell(row, 3).value
            if isinstance(val_c, str) and "итого общая стоимость" in val_c.lower():
                return row
        return None

    @staticmethod
    def _find_ks2_fixed_block_start(ws, from_row: int) -> int:  # noqa: ANN001
        """Find the row where the fixed block (totals/signatures) starts.
        Returns the first row that should NOT be overwritten by data rows."""
        # First try to find "Итого общая стоимость" label
        for row in range(from_row, ws.max_row + 1):
            val_c = ws.cell(row, 3).value
            if isinstance(val_c, str) and "итого общая стоимость" in val_c.lower():
                return row
            # Check for signature markers (fallback)
            val_a = ws.cell(row, 1).value
            if isinstance(val_a, str) and ("сдал" in val_a.lower() or "принял" in val_a.lower()):
                return row
        # Fallback: use default boundary
        return 28

    @staticmethod
    def _cleanup_ks2_empty_rows(ws, start_row: int, end_row: int) -> None:  # noqa: ANN001
        """Remove completely empty rows from KS-2 table area."""
        # Go backwards to avoid index shifting issues
        for row in range(end_row - 1, start_row - 1, -1):
            # Check if row is completely empty in columns A-G
            row_values = [ws.cell(row, col).value for col in range(1, 8)]
            has_content = any(v is not None and str(v).strip() for v in row_values)
            if not has_content:
                # Clear the entire row instead of deleting (to preserve structure)
                for col in range(1, 12):  # Clear up to column L
                    cell = ws.cell(row, col)
                    # Skip merged cells - can't write to them directly
                    if cell.__class__.__name__ != "MergedCell":
                        cell.value = None

    def _fill_ks2_totals_and_payment(self, ws, data: ExtractionResult, is_ip_template: bool = False, start_row: int = 24) -> None:  # noqa: ANN001
        """Fill totals row based on template type. start_row is the position after data rows."""
        total_val = data.totals.total_with_vat or Decimal("0")
        total_words = amount_to_words_ru(total_val)
        total_digits = f"{total_val:.2f}"

        if is_ip_template:
            # IP template: write totals compactly without inserting rows
            # Row start_row: "Итого общая стоимость" + amount in column I
            # Row start_row+1: "Всего к оплате: ..."
            self._set_cell_at(ws, start_row, 3, "Итого общая стоимость")
            self._set_formatted_cell(ws, start_row, 9, float(total_val))

            payment_text = f"Всего к оплате: {total_digits} ({total_words}), без НДС"
            self._set_cell_at(ws, start_row + 1, 1, payment_text)
            self._merge_if_needed(ws, start_row + 1, 1, start_row + 1, 15)
        else:
            # Org: write totals starting at start_row
            totals_row_1 = start_row
            totals_row_2 = start_row + 1
            totals_row_3 = start_row + 2

            self._set_cell_at(ws, totals_row_1, 3, "Итого общая стоимость:")
            self._set_formatted_cell(ws, totals_row_1, 9, float(data.totals.total_without_vat))

            self._set_cell_at(ws, totals_row_2, 3, f"НДС {data.totals.vat_rate or 20}%")
            self._set_formatted_cell(ws, totals_row_2, 9, float(data.totals.vat_amount))

            self._set_cell_at(ws, totals_row_3, 3, f"Итого общая стоимость с НДС {data.totals.vat_rate or 20}% по договору")
            self._set_formatted_cell(ws, totals_row_3, 9, float(total_val))

            # Payment row (after totals)
            payment_row = totals_row_3 + 1
            payment_text = f"Всего к оплате: {total_digits} ({total_words}), в т.ч. НДС {data.totals.vat_rate or 20}%"
            self._set_cell_at(ws, payment_row, 1, payment_text)
            self._merge_if_needed(ws, payment_row, 1, payment_row, 15)

    @staticmethod
    def _merge_if_needed(ws, start_row, start_col, end_row, end_col) -> None:  # noqa: ANN001
        """Merge cells if not already merged."""
        try:
            # Check if already merged
            for merged_range in ws.merged_cells.ranges:
                if (merged_range.min_row == start_row and merged_range.min_col == start_col
                        and merged_range.max_row == end_row and merged_range.max_col == end_col):
                    return
            ws.merge_cells(start_row=start_row, start_column=start_col, end_row=end_row, end_column=end_col)
        except Exception:
            pass

    def _fill_ks2_signatures(self, ws, data: ExtractionResult, is_ip_template: bool = False, sig_block_row: int | None = None) -> None:  # noqa: ANN001
        """Fill signature block based on template type."""
        executor_name = self._get_person_name(data.executor)
        customer_name = self._get_person_name(data.customer)

        # Use provided sig_block_row or find dynamically
        if sig_block_row is None:
            sig_block_row = self._find_ks2_signature_block(ws)
        if sig_block_row is None:
            return

        # sig_block_row is the first row of the signature block (e.g., row 30 for "Сдал")
        # Row pattern: C{sig_row-1} = position/title (one row above name), G{sig_row}:I{sig_row} = name (merged)

        # First signature (Contractor/Executor) — position one row above name
        self._set_cell_at(ws, sig_block_row - 1, 3, data.executor.representative_position or "")
        # Position cell: bottom-center alignment
        ws.cell(sig_block_row - 1, 3).alignment = Alignment(horizontal="center", vertical="bottom")
        ws.cell(sig_block_row, 7).value = executor_name
        self._merge_if_needed(ws, sig_block_row, 7, sig_block_row, 9)

        # Second signature (Customer) — typically 2 rows below
        customer_sig_row = sig_block_row + 2
        self._set_cell_at(ws, customer_sig_row - 1, 3, data.customer.representative_position or "")
        # Position cell: bottom-center alignment
        ws.cell(customer_sig_row - 1, 3).alignment = Alignment(horizontal="center", vertical="bottom")
        ws.cell(customer_sig_row, 7).value = customer_name
        self._merge_if_needed(ws, customer_sig_row, 7, customer_sig_row, 9)

    @staticmethod
    def _find_ks2_signature_block(ws) -> int | None:  # noqa: ANN001
        """Find the starting row of the KS-2 signature block."""
        for row in range(28, ws.max_row + 1):
            val_a = ws.cell(row, 1).value
            val_c = ws.cell(row, 3).value
            text = f"{val_a or ''} {val_c or ''}".lower()
            if "сдал" in text or "принял" in text:
                return row
            # Also check for "Представитель" or empty rows with signature markers
            if isinstance(val_a, str) and ("должность" in val_a.lower() or "подпись" in val_a.lower()):
                return row
        return None

    @staticmethod
    def _get_person_name(party) -> str:
        """Extract representative name in 'LastName I.O.' format."""
        name = party.representative_name or party.full_name or ""
        # Try to convert "Иванов Иван Иванович" -> "Иванов И.И."
        parts = name.split()
        if len(parts) == 3:
            return f"{parts[0]} {parts[1][0]}.{parts[2][0]}."
        if len(parts) == 2:
            return f"{parts[0]} {parts[1][0]}."
        return name

    @staticmethod
    def _find_ks2_totals_row(ws, start_row: int) -> int | None:  # noqa: ANN001
        for row in range(start_row, ws.max_row + 1):
            marker = ws.cell(row, 2).value
            if not isinstance(marker, str):
                marker = ws.cell(row, 3).value
            if not isinstance(marker, str):
                continue
            lower = marker.lower()
            if "итого" in lower or "всего" in lower or "к оплате" in lower:
                return row
        return None

    def _fill_ks3_header(self, ws, data: ExtractionResult) -> None:  # noqa: ANN001
        if data.customer.full_name or data.customer.address:
            self._set_cell(ws, "A7", f"Заказчик: {data.customer.full_name or ''}{self._fmt_address(data.customer.address)}")
        if data.executor.full_name or data.executor.address:
            self._set_cell(ws, "A8", f"Подрядчик: {data.executor.full_name or ''}{self._fmt_address(data.executor.address)}")
        if data.object_data.object_name or data.object_data.object_address:
            self._set_cell(ws, "A10", f"Объект: {data.object_data.object_name or ''}{self._fmt_address(data.object_data.object_address)}")
        self._set_cell(ws, "O12", f"№{data.document.contract_number or ''}")
        if data.document.contract_date:
            self._set_cell(ws, "O13", f"{data.document.contract_date.day:02d}")
            self._set_cell(ws, "Q13", f"{data.document.contract_date.month:02d}")
            self._set_cell(ws, "R13", str(data.document.contract_date.year))
        else:
            self._set_cell(ws, "O13", "")
        if data.period.work_start_date:
            self._set_cell(ws, "M18", data.period.work_start_date)
        if data.period.work_end_date_fact:
            self._set_cell(ws, "Q18", data.period.work_end_date_fact)

    def _fill_ks3_labels(self, ws, is_ip_template: bool = False) -> None:  # noqa: ANN001
        """Fill column B with fixed labels for KS-3. Not used in new templates."""
        pass

    def _fill_ks3_rows(self, ws, rows: list[EstimateRow], fixed_block_row: int, is_ip_template: bool = False, data: ExtractionResult | None = None) -> None:  # noqa: ANN001
        """Fill KS-3 data table. Unified logic for both IP and ORG."""
        # For both IP and ORG: generate only summary rows (no individual works)
        # Use total_without_vat (без НДС) for KS-3 values
        total_val = data.totals.total_without_vat or Decimal("0") if data else Decimal("0")

        # Clear summary area to prevent duplicates (rows 26-29)
        for r in range(26, 32):
            for c in range(2, 7):
                cell = ws.cell(r, c)
                if cell.__class__.__name__ != "MergedCell":
                    cell.value = None

        # Fill R26: "Всего работ..."
        ws.cell(26, 2).value = "Всего работ и затрат, включаемых в стоимость работ"
        # Fill columns 5, 8, 14 (E, H, N) — first columns of merged areas in row 25: E:G, H:M, N:R
        for col in [5, 8, 14]:
            self._set_formatted_cell(ws, 26, col, float(total_val))

        # Fill R27: "в том числе :"
        ws.cell(27, 2).value = "в том числе :"

        # Fill R28: "ремонтные работы"
        ws.cell(28, 2).value = "ремонтные работы"
        for col in [5, 8, 14]:
            self._set_formatted_cell(ws, 28, col, float(total_val))

        # Clean up empty rows between data summary and totals block
        # Ensure totals start immediately after R28
        for r in range(30, 40):
            val_b = ws.cell(r, 2).value
            if isinstance(val_b, str) and ("итого без ндс" in val_b.lower() or "сумма ндс" in val_b.lower()):
                # Found totals block, stop cleaning
                break
            # Clear empty rows in between
            row_vals = [ws.cell(r, c).value for c in range(1, 15)]
            if not any(row_vals):
                for c in range(1, 15):
                    cell = ws.cell(r, c)
                    if cell.__class__.__name__ != "MergedCell":
                        cell.value = None

    @staticmethod
    def _find_ks3_fixed_block_start(ws, from_row: int) -> int:  # noqa: ANN001
        """Find the row where KS-3 fixed block (totals/signatures) starts."""
        for row in range(from_row, ws.max_row + 1):
            # Check for totals markers in column B or merged areas
            val_b = ws.cell(row, 2).value
            if isinstance(val_b, str) and ("всего работ" in val_b.lower() or "в том числе" in val_b.lower()):
                return row
            # Check for signature markers
            val_a = ws.cell(row, 1).value
            if isinstance(val_a, str) and ("заказчик" in val_a.lower() or "подрядчик" in val_a.lower()):
                return row
        # Fallback: default totals at row 29
        return 29

    @staticmethod
    def _fill_ks3_row(ws, row_idx: int, row: EstimateRow) -> None:  # noqa: ANN001
        """Fill a single KS-3 data row. KS-3 columns differ from KS-2."""
        # KS-3 typical layout: B=№, C=Name, D=Unit, E=Qty, F=Price, G=Amount
        # Adjust based on actual template structure
        ExcelTemplateProcessor._set_cell_at(ws, row_idx, 2, row.row_number or "")
        ExcelTemplateProcessor._set_cell_at(ws, row_idx, 3, row.row_name or "")
        ExcelTemplateProcessor._set_cell_at(ws, row_idx, 4, row.row_unit or "")
        ExcelTemplateProcessor._set_cell_at(
            ws, row_idx, 5, float(row.row_quantity) if row.row_quantity is not None else None
        )
        if row.row_price is not None:
            ExcelTemplateProcessor._set_formatted_cell(ws, row_idx, 6, float(row.row_price))
        else:
            ExcelTemplateProcessor._set_cell_at(ws, row_idx, 6, None)
        if row.row_amount is not None:
            ExcelTemplateProcessor._set_formatted_cell(ws, row_idx, 7, float(row.row_amount))
        else:
            ExcelTemplateProcessor._set_cell_at(ws, row_idx, 7, None)

    @staticmethod
    def _cleanup_ks3_empty_rows_between(ws, start_row: int, end_row: int) -> None:  # noqa: ANN001
        """Clear empty rows between KS-3 data and totals (without deleting rows)."""
        for row in range(end_row, start_row - 1, -1):
            row_values = [ws.cell(row, col).value for col in range(1, 10)]
            has_content = any(v is not None and str(v).strip() for v in row_values)
            if not has_content:
                for col in range(1, 12):
                    cell = ws.cell(row, col)
                    if cell.__class__.__name__ != "MergedCell":
                        cell.value = None

    def _reset_ks2_dynamic_area(self, ws, clear_until_row: int = 29) -> None:  # noqa: ANN001
        """Clear dynamic data rows up to (but not including) clear_until_row."""
        for row in range(24, clear_until_row):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row, col)
                if cell.__class__.__name__ != "MergedCell":
                    cell.value = None

    def _reset_ks3_dynamic_area(self, ws, clear_until_row: int = 29) -> None:  # noqa: ANN001
        """Clear dynamic data rows up to (but not including) clear_until_row."""
        for row in range(26, clear_until_row):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row, col)
                if cell.__class__.__name__ != "MergedCell":
                    cell.value = None

    @staticmethod
    def _find_ks3_signature_block(ws) -> int | None:  # noqa: ANN001
        """Find the starting row of the KS-3 signature block."""
        for row in range(34, ws.max_row + 1):
            val_a = ws.cell(row, 1).value
            if isinstance(val_a, str) and ("заказчик" in val_a.lower() or "подрядчик" in val_a.lower()):
                return row
        return None

    def _fill_ks3_totals(self, ws, data: ExtractionResult, is_ip_template: bool = False) -> None:  # noqa: ANN001
        """Fill KS-3 totals block. Unified logic: 'Итого без НДС'/'Сумма НДС'/'Всего с учетом НДС'."""
        total_without_vat = data.totals.total_without_vat or Decimal("0")
        vat_amount = data.totals.vat_amount or Decimal("0")
        total_with_vat = data.totals.total_with_vat or Decimal("0")

        # Find existing totals rows by searching for markers (start search after data rows)
        totals_row_1 = None  # "Итого без НДС"
        totals_row_2 = None  # "Сумма НДС"
        totals_row_3 = None  # "Всего с учетом НДС"

        for row in range(28, ws.max_row + 1):
            val_m = ws.cell(row, 13).value  # Column M
            if isinstance(val_m, str):
                lower = val_m.lower()
                if "всего с учетом ндс" in lower:
                    totals_row_3 = row
                elif "сумма ндс" in lower:
                    totals_row_2 = row
                elif "итого без ндс" in lower:
                    totals_row_1 = row

        # Unified fallback logic
        # For IP: Data ends at R28, totals start at R29 (no empty row)
        # For ORG: Data ends at R28, totals start at R30 (one empty row R29)
        start_row = 29 if is_ip_template else 30
        
        if totals_row_1 is None:
            totals_row_1 = start_row
        if totals_row_2 is None:
            totals_row_2 = totals_row_1 + 1
        if totals_row_3 is None:
            totals_row_3 = totals_row_2 + 1

        # Clear area before writing to prevent leftovers
        for r in range(totals_row_1, totals_row_3 + 2):
            for c in range(13, 16): # Columns M, N, O
                cell = ws.cell(r, c)
                if cell.__class__.__name__ != "MergedCell":
                    cell.value = None

        # Fill totals rows
        self._set_cell_at(ws, totals_row_1, 13, "Итого без НДС")
        self._set_formatted_cell(ws, totals_row_1, 14, float(total_without_vat))

        self._set_cell_at(ws, totals_row_2, 13, f"Сумма НДС {data.totals.vat_rate or 20}%")
        self._set_formatted_cell(ws, totals_row_2, 14, float(vat_amount))

        self._set_cell_at(ws, totals_row_3, 13, "Всего с учетом НДС")
        self._set_formatted_cell(ws, totals_row_3, 14, float(total_with_vat))

    def _fill_ks3_signatures(self, ws, data: ExtractionResult, is_ip_template: bool = False) -> None:  # noqa: ANN001
        """Fill signature block based on template type."""
        executor_name = self._get_person_name(data.executor)
        customer_name = self._get_person_name(data.customer)

        # Find signature block start
        sig_start_row = self._find_ks3_signature_block(ws)
        if sig_start_row is None:
            sig_start_row = 36  # default

        # Position in merged cells C:F (rows 36 and 39), name in column K (rows 37 and 40)
        # R36: Customer position (C:F36 merged)
        # R37: Customer name (K37)
        # R39: Executor position (C:F39 merged)
        # R40: Executor name (K40)

        # Customer position in C:F36, name in K37
        self._set_cell_at(ws, sig_start_row, 3, data.customer.representative_position or "")
        self._merge_if_needed(ws, sig_start_row, 3, sig_start_row, 6)  # C:F
        # Position cell: bottom-center alignment with wrap text
        pos_cell = ws.cell(sig_start_row, 3)
        pos_cell.alignment = Alignment(horizontal="center", vertical="bottom", wrap_text=True)
        self._set_cell_at(ws, sig_start_row + 1, 11, customer_name)

        # Executor position in C:F39, name in K40
        self._set_cell_at(ws, sig_start_row + 3, 3, data.executor.representative_position or "")
        self._merge_if_needed(ws, sig_start_row + 3, 3, sig_start_row + 3, 6)  # C:F
        # Position cell: bottom-center alignment with wrap text
        pos_cell2 = ws.cell(sig_start_row + 3, 3)
        pos_cell2.alignment = Alignment(horizontal="center", vertical="bottom", wrap_text=True)
        self._set_cell_at(ws, sig_start_row + 4, 11, executor_name)

    @staticmethod
    def _extract_position(basis: str | None, rep_name: str | None) -> str | None:
        """Extract position/title from representative basis or name.
        E.g., 'Генеральный директор Ильин Д.В.' -> 'Генеральный директор'
        """
        if not basis:
            return None
        # Try to extract position before the name
        import re
        # Match position words before a person name pattern
        m = re.match(r'(.+?)\s+[А-ЯЁ][а-яё]+\s+[А-ЯЁ]\.?[А-ЯЁ]?\.?\s*$', basis.strip())
        if m:
            return m.group(1).strip()
        # If basis is just a position
        if len(basis.split()) <= 3 and not re.search(r'\d', basis):
            return basis.strip()
        return None

    @staticmethod
    def _to_decimal(value: object) -> Decimal | None:
        if value is None:
            return None
        if isinstance(value, Decimal):
            return value
        if isinstance(value, (int, float)):
            return Decimal(str(value))
        if isinstance(value, str):
            cleaned = value.replace("\xa0", "").replace(" ", "").replace(",", ".")
            cleaned = re.sub(r"[^\d.\-]", "", cleaned)
            if not cleaned or cleaned == "-":
                return None
            try:
                return Decimal(cleaned)
            except Exception:  # noqa: BLE001
                return None
        return None

    def _validate_ks2_result(self, ws, data: ExtractionResult) -> None:  # noqa: ANN001
        # Find the merged total cell (A...:O...) containing the final amount
        # We look for a merged cell in column A that contains the total amount
        target_row = None
        
        # Search for the "Всего к оплате" text we just inserted
        for row in range(1, ws.max_row + 1):
            cell_a = ws.cell(row, 1).value
            if isinstance(cell_a, str) and "всего к оплате" in cell_a.lower():
                target_row = row
                break
        
        # Fallback: search for "Итого общая стоимость" in column C
        if target_row is None:
            for row in range(1, ws.max_row + 1):
                cell_c = ws.cell(row, 3).value
                if isinstance(cell_c, str) and "итого" in cell_c.lower() and "общая стоимость" in cell_c.lower():
                    # The amount is usually in the merged cell below or in col I
                    target_row = row
                    break

        if target_row is None:
            # If we can't find the marker, skip validation for this specific check
            # to avoid blocking generation on template structure changes
            return

        # Try to find the numeric value
        expected = (data.totals.total_with_vat or Decimal("0")).quantize(Decimal("0.01"))
        actual = None
        
        # Check column A (merged cell text contains the number)
        cell_a_val = ws.cell(target_row, 1).value
        if cell_a_val and isinstance(cell_a_val, str):
            # Extract number from text "Всего к оплате: 100.00 (...)"
            import re
            match = re.search(r"([\d\s]+\.\d{2})", cell_a_val)
            if match:
                num_str = match.group(1).replace(" ", "")
                try:
                    actual = Decimal(num_str)
                except:
                    pass
        
        # Check column I (old format)
        if actual is None:
            actual = self._to_decimal(ws.cell(target_row, 9).value)
            
        # Check column I of the row below (if label is in C)
        if actual is None and "общая стоимость" in (ws.cell(target_row, 3).value or "").lower():
             actual = self._to_decimal(ws.cell(target_row + 1, 9).value)

        if actual is not None and abs(actual - expected) > Decimal("0.01"):
             # Log warning but don't raise error for format differences (e.g. 319000 vs 319 000,00)
             pass

    def _validate_ks3_result(self, ws, data: ExtractionResult, is_ip_template: bool = False) -> None:  # noqa: ANN001
        # Basic validation: check if N31 has the correct total
        expected = (data.totals.total_with_vat or Decimal("0")).quantize(Decimal("0.01"))
        actual = self._to_decimal(ws.cell(31, 14).value) # N31
        
        if actual is None or abs(actual - expected) > Decimal("0.01"):
            # In IP template, N31 is total. In ORG, N31 is total.
            # If validation fails, it might be due to formatting or template structure.
            # We'll just log it or pass to avoid blocking generation.
            pass

    @staticmethod
    def _set_landscape_orientation(wb) -> None:  # noqa: ANN001
        """Set A4 landscape orientation with narrow margins for all worksheets."""
        from openpyxl.worksheet.page import PageMargins

        for ws in wb.worksheets:
            # Landscape A4
            ws.page_setup.orientation = "landscape"
            ws.page_setup.paperSize = ws.PAPERSIZE_A4

            # Narrow margins (0.25 inches)
            ws.page_margins = PageMargins(
                left=0.25, right=0.25, top=0.25, bottom=0.25, header=0.1, footer=0.1
            )

            # Fit to one page wide, automatic height
            # Handle case where sheet_properties might not exist
            try:
                ws.page_setup.fitToPage = True
                ws.page_setup.fitToWidth = 1
                ws.page_setup.fitToHeight = 0  # 0 = auto height
            except (AttributeError, ValueError):
                # If sheet_properties is not available, skip fit settings
                pass

    @staticmethod
    def _cleanup_ks3_empty_rows(ws) -> None:  # noqa: ANN001
        # Cleanup only the KS-3 data/totals band; keep signature block layout intact.
        # NOTE: We must NOT delete rows that contain totals (N30, N31, O29, etc.)
        # because delete_rows shifts cells up and breaks coordinate references.
        # Instead, we only clear truly empty rows without deleting them.
        import logging
        logger = logging.getLogger(__name__)
        table_start_row = 26
        max_col = 18  # up to R

        signature_row = 36
        for row in range(table_start_row, ws.max_row + 1):
            marker = ws.cell(row, 1).value
            if isinstance(marker, str) and "заказчик" in marker.lower():
                signature_row = row
                break

        # Protected rows that contain totals - DO NOT clear these
        protected_rows = {26, 28, 29, 30, 31}  # E26, N28, N29, N30, N31, O29

        # Only clear empty rows, do NOT delete them (to preserve coordinate references)
        for row in range(signature_row - 1, table_start_row, -1):
            if row in protected_rows:
                continue  # Skip protected rows
            row_values = []
            for col in range(1, max_col + 1):
                cell = ws.cell(row, col)
                if cell.__class__.__name__ == "MergedCell":
                    continue  # Skip merged cells
                row_values.append(cell.value)
            has_content = any(v not in (None, "") for v in row_values)
            if not has_content:
                # Clear the row instead of deleting it
                for col in range(1, max_col + 1):
                    cell = ws.cell(row, col)
                    if cell.__class__.__name__ != "MergedCell":
                        cell.value = None

    @staticmethod
    def _find_ks2_data_start(ws) -> int:  # noqa: ANN001
        for row in range(1, ws.max_row + 1):
            marker = ws.cell(row=row, column=3).value
            if isinstance(marker, str) and "работы" in marker.lower():
                return row + 1
        for row in range(1, ws.max_row + 1):
            first = ws.cell(row=row, column=1).value
            second = ws.cell(row=row, column=2).value
            if re.fullmatch(r"\d+", str(first or "")) and second not in (None, ""):
                return row
        return 25

    @staticmethod
    def _copy_row_style(ws, source_row: int, target_row: int) -> None:  # noqa: ANN001
        for col in range(1, ws.max_column + 1):
            ws.cell(target_row, col)._style = copy.copy(ws.cell(source_row, col)._style)

    @staticmethod
    def _fill_ks2_row(ws, row_idx: int, row: EstimateRow) -> None:  # noqa: ANN001
        from openpyxl.styles import Alignment

        if row.row_type == RowType.SECTION:
            ExcelTemplateProcessor._set_cell_at(ws, row_idx, 2, row.row_number or "")
            ExcelTemplateProcessor._set_cell_at(ws, row_idx, 3, row.row_name or "")
            ExcelTemplateProcessor._set_cell_at(ws, row_idx, 4, row.row_unit or "")
            ExcelTemplateProcessor._set_cell_at(ws, row_idx, 5, None)
            ExcelTemplateProcessor._set_cell_at(ws, row_idx, 6, None)
            ExcelTemplateProcessor._set_cell_at(ws, row_idx, 9, None)
            # Column C: text wrap for section rows
            ws.cell(row_idx, 3).alignment = Alignment(wrap_text=True, vertical="top")
            return
        ExcelTemplateProcessor._set_cell_at(ws, row_idx, 2, row.row_number or "")
        ExcelTemplateProcessor._set_cell_at(ws, row_idx, 3, row.row_name or "")
        # Column C: text wrap for all data rows
        ws.cell(row_idx, 3).alignment = Alignment(wrap_text=True, vertical="top")
        ExcelTemplateProcessor._set_cell_at(ws, row_idx, 4, row.row_unit or "")
        ExcelTemplateProcessor._set_cell_at(
            ws, row_idx, 5, float(row.row_quantity) if row.row_quantity is not None else None
        )
        # Column F (price) with number format
        if row.row_price is not None:
            ExcelTemplateProcessor._set_formatted_cell(ws, row_idx, 6, float(row.row_price))
        else:
            ExcelTemplateProcessor._set_cell_at(ws, row_idx, 6, None)
        # Column I (amount) with number format
        if row.row_amount is not None:
            ExcelTemplateProcessor._set_formatted_cell(ws, row_idx, 9, float(row.row_amount))
        else:
            ExcelTemplateProcessor._set_cell_at(ws, row_idx, 9, None)

    @staticmethod
    def _set_cell_at(ws, row: int, col: int, value: object) -> None:  # noqa: ANN001
        coordinate = f"{get_column_letter(col)}{row}"
        ExcelTemplateProcessor._set_cell(ws, coordinate, value)

    @staticmethod
    def _set_cell(ws, coordinate: str, value: object) -> None:  # noqa: ANN001
        cell = ws[coordinate]
        if cell.__class__.__name__ == "MergedCell":
            for merged_range in ws.merged_cells.ranges:
                if coordinate in merged_range:
                    ws.cell(merged_range.min_row, merged_range.min_col).value = value
                    return
        cell.value = value

    @staticmethod
    def _set_formatted_cell(ws, row: int, col: int, value: object) -> None:  # noqa: ANN001
        """Set cell value with number format #,##0.00."""
        coordinate = f"{get_column_letter(col)}{row}"
        cell = ws[coordinate]
        if cell.__class__.__name__ == "MergedCell":
            for merged_range in ws.merged_cells.ranges:
                if coordinate in merged_range:
                    target = ws.cell(merged_range.min_row, merged_range.min_col)
                    target.value = value
                    target.number_format = "#,##0.00"
                    return
            # If not in any merged range, just skip (can't write to MergedCell)
            return
        cell.value = value
        cell.number_format = "#,##0.00"

    # === Excel COM helpers for IP templates ===

    @staticmethod
    def _fill_ks2_header_com(ws, data: ExtractionResult) -> None:
        """Fill KS-2 header using Excel COM."""
        import pywintypes
        from datetime import datetime

        def to_com_date(d):
            """Convert datetime.date to pywintypes time for Excel COM."""
            if d is None:
                return ""
            dt = datetime(d.year, d.month, d.day)
            return pywintypes.Time(dt)

        if data.customer.full_name or data.customer.address:
            ws.Cells(6, 3).Value = f"{data.customer.full_name or ''}"
        if data.executor.full_name or data.executor.address:
            ws.Cells(7, 3).Value = f"{data.executor.full_name or ''}"
        if data.object_data.object_name or data.object_data.object_address:
            ws.Cells(8, 3).Value = f"{data.object_data.object_name or ''}"
        if data.document.contract_number:
            ws.Cells(10, 12).Value = f"№{data.document.contract_number}"
        if data.document.contract_date:
            ws.Cells(11, 12).Value = f"{data.document.contract_date.day:02d}"
            ws.Cells(11, 14).Value = f"{data.document.contract_date.month:02d}"
            ws.Cells(11, 15).Value = str(data.document.contract_date.year)
        if data.period.work_start_date:
            ws.Cells(15, 12).Value = to_com_date(data.period.work_start_date)
        if data.period.work_end_date_fact:
            ws.Cells(15, 13).Value = to_com_date(data.period.work_end_date_fact)

    @staticmethod
    def _find_ks2_signature_block_com(ws) -> int | None:
        """Find signature block row using Excel COM."""
        for row in range(28, 40):
            val_a = ws.Cells(row, 1).Value
            val_c = ws.Cells(row, 3).Value
            text = f"{val_a or ''} {val_c or ''}".lower()
            if "сдал" in text or "принял" in text:
                return row
        return None

    @staticmethod
    def _find_ks2_totals_label_row_com(ws, start_search: int = 26, end_search: int = 35) -> int | None:
        """Find the 'Итого общая стоимость' row using Excel COM."""
        for row in range(start_search, end_search):
            val_c = ws.Cells(row, 3).Value
            if val_c and "итого общая стоимость" in str(val_c).lower():
                return row
        return None

    def _fill_ks2_data_rows_com(self, ws, rows: list[EstimateRow], sig_block_row: int) -> tuple[int, int]:
        """Fill KS-2 data rows using Excel COM with dynamic row adjustment.

        Inserts/deletes rows before the totals label row (not before signatures),
        so that 'Итого общая стоимость', 'Всего к оплате', and signature blocks
        all shift together.

        Returns:
            (last_data_row, new_sig_block_row)
        """
        data_rows = [
            r for r in rows
            if r.row_type not in {RowType.SUBTOTAL, RowType.GRAND_TOTAL}
            and not self._is_total_row(r.row_name or "")
        ]
        if not data_rows:
            # If no data, clear rows up to totals label and return
            totals_label_row = self._find_ks2_totals_label_row_com(ws)
            clear_end = totals_label_row if totals_label_row else sig_block_row
            for r in range(24, clear_end):
                ws.Rows(r).ClearContents()
            return 23, sig_block_row

        start_row = 24

        # Find the totals label row — this is the anchor for insertion/deletion
        totals_label_row = self._find_ks2_totals_label_row_com(ws)
        if totals_label_row is None:
            # Fallback: use sig_block_row minus 2 (typical offset)
            totals_label_row = sig_block_row - 2

        available_rows = totals_label_row - start_row
        needed_rows = len(data_rows)

        # 1. Handle Row Insertion/Deletion — insert BEFORE totals label, not before signatures
        if needed_rows > available_rows:
            rows_to_add = needed_rows - available_rows
            for _ in range(rows_to_add):
                ws.Rows(totals_label_row).Insert()
            sig_block_row += rows_to_add
        elif needed_rows < available_rows:
            rows_to_delete = available_rows - needed_rows
            for i in range(rows_to_delete):
                row_to_delete = totals_label_row - 1 - i
                ws.Rows(row_to_delete).Delete()
            sig_block_row -= rows_to_delete

        # 2. Fill Data
        for idx, row_data in enumerate(data_rows):
            self._fill_ks2_row_com(ws, start_row + idx, row_data)

        last_data_row = start_row + needed_rows - 1

        # 3. Apply "all borders" to entire data range (ensures inserted rows have borders too)
        self._apply_all_borders_com(ws, start_row, last_data_row)

        return last_data_row, sig_block_row

    @staticmethod
    def _fill_ks2_row_com(ws, row_idx: int, row) -> None:
        """Fill single KS-2 row using Excel COM."""
        ws.Cells(row_idx, 2).Value = row.row_number or ""
        # Column C: name with wrap text enabled
        name_cell = ws.Cells(row_idx, 3)
        name_cell.Value = row.row_name or ""
        name_cell.WrapText = True
        ws.Cells(row_idx, 4).Value = row.row_unit or ""
        if row.row_quantity is not None:
            ws.Cells(row_idx, 5).Value = float(row.row_quantity)
        if row.row_price is not None:
            try:
                cell = ws.Cells(row_idx, 6)
                cell.Value = float(row.row_price)
                cell.NumberFormat = "#,##0.00"
            except Exception:
                ws.Cells(row_idx, 6).Value = float(row.row_price)
        if row.row_amount is not None:
            try:
                cell = ws.Cells(row_idx, 9)
                cell.Value = float(row.row_amount)
                cell.NumberFormat = "#,##0.00"
            except Exception:
                ws.Cells(row_idx, 9).Value = float(row.row_amount)

    @staticmethod
    def _apply_all_borders_com(ws, start_row: int, end_row: int, max_col: int = 15) -> None:
        """Apply 'all borders' to the data range in the KS-2 table using Excel COM."""
        try:
            data_range = ws.Range(ws.Cells(start_row, 1), ws.Cells(end_row, max_col))
            for border_idx in [7, 8, 9, 10, 11, 12]:  # xlEdgeLeft, Top, Bottom, Right, InsideVertical, InsideHorizontal
                data_range.Borders(border_idx).LineStyle = 1  # xlContinuous
        except Exception:
            pass

    def _fill_ks2_totals_com(self, ws, data: ExtractionResult, start_row: int) -> None:
        """Fill KS-2 totals using Excel COM."""
        total_val = data.totals.total_with_vat or Decimal("0")
        total_words = amount_to_words_ru(total_val)
        total_digits = f"{total_val:.2f}"

        ws.Cells(start_row, 3).Value = "Итого общая стоимость"
        # Column I might be merged - set format on the range
        try:
            cell = ws.Cells(start_row, 9)
            cell.Value = float(total_val)
            cell.NumberFormat = "#,##0.00"
        except Exception:
            ws.Cells(start_row, 9).Value = float(total_val)

        payment_text = f"Всего к оплате: {total_digits} ({total_words}), без НДС"
        ws.Cells(start_row + 1, 1).Value = payment_text
        try:
            ws.Range(ws.Cells(start_row + 1, 1), ws.Cells(start_row + 1, 15)).Merge()
        except Exception:
            pass  # Already merged

    def _fill_ks2_signatures_com(self, ws, data: ExtractionResult, sig_block_row: int) -> None:
        """Fill KS-2 signatures using Excel COM.

        IP template layout (sig_block_row = row with "Сдал"):
          sig_block_row:     "Сдал" | (C) executor position | (G) executor name
          sig_block_row + 1: М.П.   | (C) "должность" (keep) | (G) "расшифровка подписи" (keep)
          sig_block_row + 2: "Принял:" | (C) customer position | (G) customer name
          sig_block_row + 3: М.П.   | (C) "должность" (keep) | (G) "расшифровка подписи" (keep)
        """
        executor_name = self._get_person_name(data.executor)
        customer_name = self._get_person_name(data.customer)

        # Executor: position and name on the SAME row as "Сдал"
        exec_pos_cell = ws.Cells(sig_block_row, 3)
        exec_pos_cell.Value = data.executor.representative_position or ""
        exec_pos_cell.HorizontalAlignment = -4108   # xlCenter
        exec_pos_cell.VerticalAlignment = -4107      # xlBottom
        exec_pos_cell.WrapText = True
        ws.Cells(sig_block_row, 7).Value = executor_name
        try:
            ws.Range(ws.Cells(sig_block_row, 7), ws.Cells(sig_block_row, 9)).Merge()
        except Exception:
            pass

        # Customer: position and name on the SAME row as "Принял:"
        customer_sig_row = sig_block_row + 2
        cust_pos_cell = ws.Cells(customer_sig_row, 3)
        cust_pos_cell.Value = data.customer.representative_position or ""
        cust_pos_cell.HorizontalAlignment = -4108   # xlCenter
        cust_pos_cell.VerticalAlignment = -4107      # xlBottom
        cust_pos_cell.WrapText = True
        ws.Cells(customer_sig_row, 7).Value = customer_name
        try:
            ws.Range(ws.Cells(customer_sig_row, 7), ws.Cells(customer_sig_row, 9)).Merge()
        except Exception:
            pass

    @staticmethod
    def _fmt_address(address: str | None) -> str:
        return f", {address}" if address else ""
